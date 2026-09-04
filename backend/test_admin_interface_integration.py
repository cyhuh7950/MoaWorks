from __future__ import annotations

import unittest
from unittest.mock import MagicMock, patch

from fastapi import HTTPException
from openpyxl import Workbook
from pydantic import ValidationError

from app.api.router import api_router
from app.api.routes.admin import apply_org_import, create_user, update_user, validate_org_import
from app.schemas.content_operations import ContentBulkStatus
from app.schemas.directory import (
    AuthUserSummary,
    DepartmentUpdateRequest,
    OrgImportApplyRequest,
    UserCreateRequest,
    UserUpdateRequest,
)
from app.services.directory_store import DirectoryAdminActiveLimitError, DirectoryStore
from app.services.org_import_service import OrgImportService


class AdminInterfaceIntegrationContractTest(unittest.TestCase):
    def test_user_create_and_update_map_active_limit_to_stable_409(self) -> None:
        actor = MagicMock(spec=AuthUserSummary)
        create_payload = UserCreateRequest(
            name="QA 관리자",
            loginId="qa.admin",
            password="safe-pass-123",
            departmentId="department-1",
            roleId="role-1",
            userType="admin",
        )
        with patch.object(
            DirectoryStore,
            "create_user",
            side_effect=DirectoryAdminActiveLimitError("관리자 활성 계정은 최대 3개입니다."),
        ):
            with self.assertRaises(HTTPException) as create_context:
                create_user(payload=create_payload, _=actor)

        with patch.object(
            DirectoryStore,
            "update_user",
            side_effect=DirectoryAdminActiveLimitError("관리자 활성 계정은 최대 3개입니다."),
        ):
            with self.assertRaises(HTTPException) as update_context:
                update_user(user_id="user-4", payload=UserUpdateRequest(userType="admin"), _=actor)

        for context in (create_context, update_context):
            self.assertEqual(context.exception.status_code, 409)
            self.assertEqual(context.exception.detail["code"], "ADMIN_ACTIVE_LIMIT_REACHED")

    def test_org_import_apply_maps_active_limit_to_stable_409(self) -> None:
        actor = MagicMock(spec=AuthUserSummary)
        payload = OrgImportApplyRequest(batchId="batch-admin-limit")
        with patch.object(
            OrgImportService,
            "apply_batch",
            side_effect=DirectoryAdminActiveLimitError("활성 관리자 계정은 최대 3개까지 사용할 수 있습니다."),
        ):
            with self.assertRaises(HTTPException) as context:
                apply_org_import(payload=payload, actor=actor)

        self.assertEqual(context.exception.status_code, 409)
        self.assertEqual(context.exception.detail["code"], "ADMIN_ACTIVE_LIMIT_REACHED")

    def test_required_admin_routes_are_registered(self) -> None:
        registered = {(route.path, method) for route in api_router.routes for method in getattr(route, "methods", set())}
        expected = {
            ("/admin/departments/{department_id}", "PATCH"),
            ("/admin/departments/{department_id}", "DELETE"),
            ("/admin/roles/{role_id}", "DELETE"),
            ("/admin/users/{user_id}", "DELETE"),
            ("/admin/org-import/template", "GET"),
            ("/admin/org-import/validate", "POST"),
            ("/admin/org-import/apply", "POST"),
            ("/admin/content/messages", "GET"),
            ("/admin/content/help-policies", "GET"),
            ("/admin/mail-delivery/status", "GET"),
            ("/admin/mail-delivery/provider", "PATCH"),
            ("/admin/mail-operations/submission-credentials", "GET"),
            ("/admin/mail-operations/submission-credentials/{user_id}/issue", "POST"),
            ("/admin/mail-operations/submission-credentials/{user_id}/revoke", "POST"),
        }
        self.assertTrue(expected.issubset(registered), expected - registered)

    def test_user_creation_requires_safe_explicit_password(self) -> None:
        with self.assertRaises(ValidationError):
            UserCreateRequest(name="테스트", loginId="test.user", password="short", departmentId="d", roleId="r")
        request = UserCreateRequest(
            name="테스트", loginId="test.user", password="safe-pass-123", departmentId="d", roleId="r"
        )
        self.assertEqual(request.loginId, "test.user")

    def test_update_and_apply_inputs_reject_invalid_values(self) -> None:
        with self.assertRaises(ValidationError):
            DepartmentUpdateRequest(status="deleted")
        with self.assertRaises(ValidationError):
            OrgImportApplyRequest(batchId="")
        with self.assertRaises(ValidationError):
            ContentBulkStatus(ids=["message-1"], status="unexpected")

    def test_org_import_template_contains_required_sheets(self) -> None:
        service = OrgImportService()
        workbook = service.build_template()
        self.assertGreater(len(workbook), 100)
        departments, users = service._parse_workbook(workbook)
        self.assertEqual(len(departments), 3)
        self.assertEqual(len(users), 2)

    def test_org_import_rejects_unsupported_extension_and_mime(self) -> None:
        service = OrgImportService()
        with self.assertRaisesRegex(ValueError, "xlsx"):
            service.validate_file_metadata("organization.xls", service.XLSX_MIME_TYPE)
        with self.assertRaisesRegex(ValueError, "형식"):
            service.validate_file_metadata("organization.xlsx", "text/plain")

    def test_org_import_rejects_excessive_zip_resources(self) -> None:
        service = OrgImportService()
        too_many_entries = [MagicMock(file_size=1) for _ in range(service.MAX_ZIP_ENTRIES + 1)]
        excessive_size_entries = [MagicMock(file_size=service.MAX_UNCOMPRESSED_BYTES + 1)]

        for entries in (too_many_entries, excessive_size_entries):
            archive = MagicMock()
            archive.__enter__.return_value.infolist.return_value = entries
            with patch("app.services.org_import_service.ZipFile", return_value=archive):
                with self.assertRaises(ValueError):
                    service.validate_workbook_archive(b"PK")

    def test_org_import_rejects_more_than_max_data_rows(self) -> None:
        service = OrgImportService()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "departments"
        worksheet.append(service.DEPARTMENT_HEADERS)
        worksheet.cell(row=service.MAX_DATA_ROWS + 2, column=1, value="OVER")
        with self.assertRaisesRegex(ValueError, "10,000"):
            service._read_sheet_rows(worksheet, service.DEPARTMENT_HEADERS)

    def test_org_import_parse_error_does_not_expose_internal_exception(self) -> None:
        service = OrgImportService()
        with (
            patch.object(service, "validate_workbook_archive"),
            patch("app.services.org_import_service.load_workbook", side_effect=RuntimeError("C:\\secret\\internal.xlsx")),
        ):
            with self.assertRaises(ValueError) as context:
                service._parse_workbook(b"invalid")
        self.assertNotIn("secret", str(context.exception))


class AdminOrgImportUploadLimitTest(unittest.IsolatedAsyncioTestCase):
    async def test_route_reads_only_limit_plus_one_and_rejects_oversize(self) -> None:
        class OversizeUpload:
            filename = "organization.xlsx"
            content_type = OrgImportService.XLSX_MIME_TYPE

            def __init__(self) -> None:
                self.requested_size = 0

            async def read(self, size: int = -1) -> bytes:
                self.requested_size = size
                return b"x" * size

        upload = OversizeUpload()
        actor = AuthUserSummary(
            userId="admin-user",
            userName="관리자",
            userEmail="admin@moaworks.local",
            companyId="company",
            roleId="role",
            roleName="관리자",
            permissions=["admin"],
            userType="admin",
            status="active",
        )
        with self.assertRaises(HTTPException) as context:
            await validate_org_import(file=upload, deactivation_scope="uploaded_departments_only", actor=actor)
        self.assertEqual(upload.requested_size, OrgImportService.MAX_UPLOAD_BYTES + 1)
        self.assertEqual(context.exception.status_code, 413)


if __name__ == "__main__":
    unittest.main()

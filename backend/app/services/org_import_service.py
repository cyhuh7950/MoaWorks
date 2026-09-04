from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
import json
from pathlib import Path
from uuid import uuid4
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook

from fastapi import HTTPException, status

from app.schemas.directory import (
    AuthUserSummary,
    OrgImportApplyRequest,
    OrgImportBatchResponse,
    OrgImportDeactivationPreview,
    OrgImportDepartmentPreview,
    OrgImportIssue,
    OrgImportUserPreview,
)
from app.services.postgres_service import PostgresService
from app.services.security_service import SecurityService
from app.services.directory_store import _map_admin_active_limit


class OrgImportService:
    MAX_UPLOAD_BYTES = 10 * 1024 * 1024
    MAX_ZIP_ENTRIES = 128
    MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
    MAX_DATA_ROWS = 10_000
    XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ALLOWED_MIME_TYPES = frozenset({XLSX_MIME_TYPE, "application/octet-stream"})
    DEPARTMENT_HEADERS = ["department_code", "department_name", "parent_department_code", "sort_order", "status"]
    USER_HEADERS = ["login_id", "name", "department_code", "role_code", "status"]
    DEFAULT_DEACTIVATION_SCOPE = "uploaded_departments_only"
    COMPANY_ALL_CONFIRMATION_TEXT = "회사 전체 누락 사용자 비활성화에 동의합니다."
    DEACTIVATION_CONFIRMATION_TEXT = "누락 사용자 비활성화에 동의합니다."
    VERIFICATION_ACCOUNT_PREFIXES = ("verify.", "test.", "incident.", "bulk.", "layoutverify", "roleverify")
    EXPLICIT_PROTECTED_EMAILS = {"admin@moaworks.local", "cyhuh@moaworks.local", "ysla@moaworks.local"}

    def __init__(self) -> None:
        self.db = PostgresService()
        self.security = SecurityService()

    def build_template(self) -> bytes:
        workbook = Workbook()
        department_sheet = workbook.active
        department_sheet.title = "departments"
        department_sheet.append(self.DEPARTMENT_HEADERS)
        department_sheet.append(["HQ", "본사", "", 100, "active"])
        department_sheet.append(["SALES", "영업팀", "HQ", 110, "active"])
        department_sheet.append(["OPS", "운영팀", "HQ", 120, "active"])

        user_sheet = workbook.create_sheet("users")
        user_sheet.append(self.USER_HEADERS)
        user_sheet.append(["hong.gildong", "홍길동", "SALES", "일반사용자", "active"])
        user_sheet.append(["lee.sun", "이순신", "OPS", "일반사용자", "active"])

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def validate_file_metadata(self, file_name: str, content_type: str | None) -> None:
        if Path(file_name or "").suffix.lower() != ".xlsx":
            raise ValueError("조직 일괄 등록은 .xlsx 파일만 지원합니다.")
        normalized_content_type = str(content_type or "").split(";", 1)[0].strip().lower()
        if normalized_content_type not in self.ALLOWED_MIME_TYPES:
            raise ValueError("지원하지 않는 엑셀 파일 형식입니다.")

    def validate_workbook_archive(self, content: bytes) -> None:
        if len(content) > self.MAX_UPLOAD_BYTES:
            raise ValueError("업로드 파일은 10 MiB 이하여야 합니다.")
        try:
            with ZipFile(BytesIO(content)) as archive:
                entries = archive.infolist()
                if len(entries) > self.MAX_ZIP_ENTRIES:
                    raise ValueError("엑셀 내부 파일 항목 수가 허용 범위를 초과했습니다.")
                uncompressed_bytes = sum(max(0, int(entry.file_size)) for entry in entries)
                if uncompressed_bytes > self.MAX_UNCOMPRESSED_BYTES:
                    raise ValueError("엑셀 압축 해제 예상 크기가 허용 범위를 초과했습니다.")
        except BadZipFile as exc:
            raise ValueError("올바른 .xlsx 파일이 아닙니다.") from exc

    def validate_upload(self, actor: AuthUserSummary, file_name: str, content: bytes, deactivation_scope: str | None = None) -> OrgImportBatchResponse:
        self.db.ensure_migrations_applied()
        departments, users = self._parse_workbook(content)
        with self.db.connect() as connection:
            with connection.cursor() as cursor:
                analysis = self._analyze_import(cursor, departments, users, actor=actor, deactivation_scope=deactivation_scope)
                batch_id = self._new_id("orgimport")
                now = self._now()
                preview_json = self._analysis_to_preview_json(analysis)
                payload_json = {
                    "departments": analysis["departments"],
                    "users": analysis["users"],
                    "deactivationScope": analysis["deactivation_scope"],
                }
                deactivation_preview = self._deactivation_preview_rows(analysis["users_to_deactivate"])
                cursor.execute(
                    """
                    INSERT INTO org_import_batches (
                        id, company_id, uploaded_by_user_id, uploaded_by_user_name, file_name,
                        validation_status, apply_status,
                        inactive_department_count, created_department_count, moved_user_count, created_user_count, deactivated_user_count,
                        errors_json, warnings_json, preview_json, payload_json, deactivated_users_json, uploaded_at
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s::jsonb, %s::jsonb, %s::jsonb, %s::jsonb, %s)
                    RETURNING *
                    """,
                    (
                        batch_id,
                        analysis["company_id"],
                        actor.userId,
                        actor.userName,
                        file_name,
                        "passed" if not analysis["errors"] else "failed",
                        "pending",
                        analysis["inactive_department_count"],
                        analysis["created_department_count"],
                        analysis["moved_user_count"],
                        analysis["created_user_count"],
                        analysis["deactivated_user_count"],
                        json.dumps(analysis["errors"], ensure_ascii=False),
                        json.dumps(analysis["warnings"], ensure_ascii=False),
                        json.dumps(preview_json, ensure_ascii=False),
                        json.dumps(payload_json, ensure_ascii=False),
                        json.dumps(deactivation_preview, ensure_ascii=False),
                        now,
                    ),
                )
                row = cursor.fetchone()
            connection.commit()
        return self._row_to_batch_response(row)

    @_map_admin_active_limit
    def apply_batch(self, actor: AuthUserSummary, payload: OrgImportApplyRequest) -> OrgImportBatchResponse:
        self.db.ensure_migrations_applied()
        with self.db.connect() as connection:
            with connection.cursor() as cursor:
                batch = self._fetch_batch_row(cursor, payload.batchId, for_update=True)
                if batch["apply_status"] == "applied":
                    raise ValueError("이미 적용이 완료된 배치입니다.")

                payload_json = self._json_value(batch["payload_json"], default={})
                deactivation_scope = self._normalize_deactivation_scope(payload_json.get("deactivationScope"))
                analysis = self._analyze_import(
                    cursor,
                    payload_json.get("departments", []),
                    payload_json.get("users", []),
                    actor=actor,
                    deactivation_scope=deactivation_scope,
                    preserve_system_codes=True,
                )
                if analysis["errors"]:
                    raise ValueError("검증 오류가 남아 있어 적용할 수 없습니다.")

                deactivation_preview = self._deactivation_preview_rows(analysis["users_to_deactivate"])
                protected_preview = self._deactivation_preview_rows(analysis["protected_users"])
                stored_preview = self._json_value(batch["preview_json"], default={})
                stored_deactivation = self._json_value(batch.get("deactivated_users_json"), default=stored_preview.get("usersToDeactivate", []))
                stored_protected = stored_preview.get("protectedUsers", [])

                if not self._same_preview_targets(stored_deactivation, deactivation_preview) or not self._same_preview_targets(stored_protected, protected_preview):
                    raise HTTPException(
                        status_code=status.HTTP_409_CONFLICT,
                        detail={
                            "code": "ORG_IMPORT_REVALIDATION_REQUIRED",
                            "userMessage": "검증 이후 대상 목록이 변경되어 다시 검증해야 합니다.",
                            "adminMessage": f"org import apply blocked: preview changed for batch {payload.batchId}",
                        },
                    )

                if protected_preview:
                    raise HTTPException(
                        status_code=status.HTTP_409_CONFLICT,
                        detail={
                            "code": "ORG_IMPORT_PROTECTED_USERS_BLOCKED",
                            "userMessage": "보호 계정이 누락 사용자 비교에 포함되어 적용할 수 없습니다. 업로드 파일을 수정한 뒤 다시 검증하세요.",
                            "adminMessage": f"org import apply blocked: protected users present for batch {payload.batchId}",
                        },
                    )

                if deactivation_scope == "company_all":
                    confirmed = (payload.confirmationText or "").strip() == self.COMPANY_ALL_CONFIRMATION_TEXT
                else:
                    confirmed = payload.confirmDeactivateMissingUsers or (payload.confirmationText or "").strip() == self.DEACTIVATION_CONFIRMATION_TEXT
                if deactivation_preview and not confirmed:
                    raise HTTPException(
                        status_code=status.HTTP_409_CONFLICT,
                        detail={
                            "code": "ORG_IMPORT_DEACTIVATION_CONFIRM_REQUIRED",
                            "userMessage": "누락 사용자 비활성화 대상 확인이 필요합니다.",
                            "adminMessage": f"org import apply blocked: deactivation confirmation required for batch {payload.batchId}",
                        },
                    )

                now = self._now()
                created_department_ids: dict[int, str] = {}
                for department in analysis["current_active_departments"]:
                    cursor.execute(
                        "UPDATE departments SET status = %s WHERE id = %s",
                        ("inactive", department["id"]),
                    )
                    self._insert_audit(
                        cursor=cursor,
                        company_id=analysis["company_id"],
                        actor_user_id=actor.userId,
                        actor_user_name=actor.userName,
                        target_type="department",
                        target_id=department["id"],
                        event="directory.department_updated",
                        status_before=department["status"],
                        status_after="inactive",
                        reason="조직/사용자 일괄 업로드 기존 부서 비활성화",
                    )

                for department in analysis["departments"]:
                    department_id = self._new_id("dept")
                    created_department_ids[department["rowNumber"]] = department_id
                    parent_id = created_department_ids.get(department["parentRowNumber"]) if department.get("parentRowNumber") else None
                    cursor.execute(
                        """
                        INSERT INTO departments (
                            id, company_id, system_department_code, department_code, name, parent_id, status, sort_order, created_at
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        (
                            department_id,
                            analysis["company_id"],
                            department["systemDepartmentCode"],
                            department["departmentCode"],
                            department["departmentName"],
                            parent_id,
                            department["status"],
                            department["sortOrder"],
                            now,
                        ),
                    )
                    self._insert_audit(
                        cursor=cursor,
                        company_id=analysis["company_id"],
                        actor_user_id=actor.userId,
                        actor_user_name=actor.userName,
                        target_type="department",
                        target_id=department_id,
                        event="directory.department_created",
                        status_before=None,
                        status_after=department["status"],
                        reason="조직/사용자 일괄 업로드 신규 부서 생성",
                    )

                for user in analysis["users"]:
                    department_id = created_department_ids[user["departmentRowNumber"]]
                    if user["existingUserId"]:
                        cursor.execute(
                            """
                            UPDATE users
                            SET name = %s,
                                department_id = %s,
                                role_id = %s,
                                status = %s,
                                updated_at = %s
                            WHERE id = %s
                            """,
                            (
                                user["name"],
                                department_id,
                                user["roleId"],
                                user["status"],
                                now,
                                user["existingUserId"],
                            ),
                        )
                        cursor.execute(
                            """
                            UPDATE mail_accounts
                            SET status = %s,
                                updated_at = %s
                            WHERE user_id = %s
                            """,
                            ("active" if user["status"] == "active" else "inactive", now, user["existingUserId"]),
                        )
                        self._insert_audit(
                            cursor=cursor,
                            company_id=analysis["company_id"],
                            actor_user_id=actor.userId,
                            actor_user_name=actor.userName,
                            target_type="user",
                            target_id=user["existingUserId"],
                            event="directory.user_updated",
                            status_before=user["existingStatus"],
                            status_after=user["status"],
                            reason="조직/사용자 일괄 업로드 사용자 재배치",
                        )
                    else:
                        user_id = self._new_id("user")
                        mail_account_id = self._new_id("mail")
                        email = f"{user['loginId']}@{analysis['company_domain']}"
                        cursor.execute(
                            """
                            INSERT INTO users (
                                id, company_id, email, name, password_hash, department_id, role_id,
                                status, user_type, must_change_password, created_at, updated_at
                            )
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (
                                user_id,
                                analysis["company_id"],
                                email,
                                user["name"],
                                self.security.hash_password(user["loginId"]),
                                department_id,
                                user["roleId"],
                                user["status"],
                                "user",
                                True,
                                now,
                                now,
                            ),
                        )
                        cursor.execute(
                            """
                            INSERT INTO mail_accounts (
                                id, user_id, email, quota_mb, status, created_at, updated_at
                            )
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                            """,
                            (
                                mail_account_id,
                                user_id,
                                email,
                                2048,
                                "active" if user["status"] == "active" else "inactive",
                                now,
                                now,
                            ),
                        )
                        self._insert_audit(
                            cursor=cursor,
                            company_id=analysis["company_id"],
                            actor_user_id=actor.userId,
                            actor_user_name=actor.userName,
                            target_type="user",
                            target_id=user_id,
                            event="directory.user_created",
                            status_before=None,
                            status_after=user["status"],
                            reason="엑셀 일괄 업로드 생성 / 최초 로그인 후 비밀번호 변경 필요",
                        )

                for user in analysis["users_to_deactivate"]:
                    cursor.execute(
                        """
                        UPDATE users
                        SET status = 'inactive',
                            updated_at = %s
                        WHERE id = %s
                        """,
                        (now, user["id"]),
                    )
                    cursor.execute(
                        """
                        UPDATE mail_accounts
                        SET status = 'inactive',
                            updated_at = %s
                        WHERE user_id = %s
                        """,
                        (now, user["id"]),
                    )
                    self._insert_audit(
                        cursor=cursor,
                        company_id=analysis["company_id"],
                        actor_user_id=actor.userId,
                        actor_user_name=actor.userName,
                        target_type="user",
                        target_id=user["id"],
                        event="directory.user_updated",
                        status_before=user["status"],
                        status_after="inactive",
                        reason=f"업로드 파일 누락 사용자 비활성화 / batch {payload.batchId}",
                    )

                cursor.execute(
                    """
                    UPDATE org_import_batches
                    SET apply_status = %s,
                        applied_at = %s,
                        inactive_department_count = %s,
                        created_department_count = %s,
                        moved_user_count = %s,
                        created_user_count = %s,
                        deactivated_user_count = %s,
                        errors_json = %s::jsonb,
                        warnings_json = %s::jsonb,
                        preview_json = %s::jsonb,
                        payload_json = %s::jsonb,
                        deactivated_users_json = %s::jsonb
                    WHERE id = %s
                    RETURNING *
                    """,
                    (
                        "applied",
                        now,
                        analysis["inactive_department_count"],
                        analysis["created_department_count"],
                        analysis["moved_user_count"],
                        analysis["created_user_count"],
                        analysis["deactivated_user_count"],
                        json.dumps(analysis["errors"], ensure_ascii=False),
                        json.dumps(analysis["warnings"], ensure_ascii=False),
                        json.dumps(self._analysis_to_preview_json(analysis), ensure_ascii=False),
                        json.dumps({"departments": analysis["departments"], "users": analysis["users"], "deactivationScope": analysis["deactivation_scope"]}, ensure_ascii=False),
                        json.dumps(deactivation_preview, ensure_ascii=False),
                        payload.batchId,
                    ),
                )
                row = cursor.fetchone()
                self._insert_audit(
                    cursor=cursor,
                    company_id=analysis["company_id"],
                    actor_user_id=actor.userId,
                    actor_user_name=actor.userName,
                    target_type="org_import",
                    target_id=payload.batchId,
                    event="directory.org_import_applied",
                    status_before=batch["apply_status"],
                    status_after="applied",
                    reason=f"신규 부서 {analysis['created_department_count']}개 / 신규 사용자 {analysis['created_user_count']}명 / 이동 사용자 {analysis['moved_user_count']}명 / 비활성화 사용자 {analysis['deactivated_user_count']}명",
                )
            connection.commit()
        return self._row_to_batch_response(row)

    def get_batch(self, batch_id: str) -> OrgImportBatchResponse:
        self.db.ensure_migrations_applied()
        with self.db.connect() as connection:
            with connection.cursor() as cursor:
                row = self._fetch_batch_row(cursor, batch_id)
        return self._row_to_batch_response(row)

    def _parse_workbook(self, content: bytes) -> tuple[list[dict], list[dict]]:
        self.validate_workbook_archive(content)
        try:
            workbook = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001
            raise ValueError("엑셀 파일을 읽지 못했습니다. 올바른 템플릿인지 확인하세요.") from exc

        try:
            if "departments" not in workbook.sheetnames or "users" not in workbook.sheetnames:
                raise ValueError("엑셀 파일에는 departments, users 시트가 모두 있어야 합니다.")
            department_rows = self._read_sheet_rows(workbook["departments"], self.DEPARTMENT_HEADERS)
            user_rows = self._read_sheet_rows(workbook["users"], self.USER_HEADERS)
            return department_rows, user_rows
        finally:
            workbook.close()

    def _read_sheet_rows(self, worksheet, expected_headers: list[str]) -> list[dict]:
        if int(worksheet.max_row or 0) > self.MAX_DATA_ROWS + 1:
            raise ValueError(f"{worksheet.title} 시트의 데이터 행은 10,000행 이하여야 합니다.")
        header_cells = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_cells is None:
            raise ValueError(f"{worksheet.title} 시트 헤더를 찾을 수 없습니다.")
        actual_headers = [self._normalize_text(value) for value in header_cells]
        if actual_headers[: len(expected_headers)] != expected_headers:
            raise ValueError(f"{worksheet.title} 시트 헤더가 올바르지 않습니다.")

        rows: list[dict] = []
        for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row = {expected_headers[index]: values[index] if index < len(values) else None for index in range(len(expected_headers))}
            if all(self._normalize_text(value) == "" for value in row.values()):
                continue
            row["rowNumber"] = row_number
            rows.append(row)
        return rows

    def _analyze_import(
        self,
        cursor,
        departments: list[dict],
        users: list[dict],
        *,
        actor: AuthUserSummary | None = None,
        deactivation_scope: str | None = None,
        preserve_system_codes: bool = False,
    ) -> dict:
        company = self._fetch_company_row(cursor)
        cursor.execute(
            "SELECT id, company_id, system_department_code, department_code, name, parent_id, status, sort_order, created_at FROM departments ORDER BY sort_order ASC, created_at ASC"
        )
        existing_departments = cursor.fetchall()
        cursor.execute("SELECT id, company_id, name, permissions, status, created_at FROM roles ORDER BY created_at ASC")
        roles = cursor.fetchall()
        cursor.execute(
            """
            SELECT u.id, u.email, u.name, u.status, u.user_type, u.department_id, u.role_id, u.must_change_password,
                   d.name AS department_name, d.department_code, d.system_department_code, r.name AS role_name
            FROM users u
            JOIN departments d ON d.id = u.department_id
            JOIN roles r ON r.id = u.role_id
            ORDER BY u.created_at ASC
            """
        )
        existing_users = cursor.fetchall()

        errors: list[dict] = []
        warnings: list[dict] = []
        existing_system_codes = {item.get("system_department_code") for item in existing_departments if item.get("system_department_code")}
        seen_system_codes: set[str] = set(existing_system_codes)

        normalized_departments: list[dict] = []
        department_code_map: dict[str, list[dict]] = {}
        for raw in departments:
            department_code = self._normalize_text(raw.get("department_code") or raw.get("departmentCode"))
            department_name = self._normalize_text(raw.get("department_name") or raw.get("departmentName"))
            parent_department_code = self._normalize_text(raw.get("parent_department_code") or raw.get("parentDepartmentCode")) or None
            sort_order = self._normalize_int(raw.get("sort_order") if "sort_order" in raw else raw.get("sortOrder"), default=100)
            status = self._normalize_status(raw.get("status"), default="active")
            system_department_code = self._normalize_text(raw.get("systemDepartmentCode")) if preserve_system_codes else ""
            if not system_department_code:
                system_department_code = self._new_unique_system_department_code(seen_system_codes)
            else:
                seen_system_codes.add(system_department_code)
            row_number = int(raw["rowNumber"])
            if not department_code:
                errors.append(self._issue("error", "departments", row_number, "department_code는 필수입니다."))
            if not department_name:
                errors.append(self._issue("error", "departments", row_number, "department_name은 필수입니다."))
            normalized = {
                "rowNumber": row_number,
                "departmentCode": department_code,
                "departmentName": department_name,
                "parentDepartmentCode": parent_department_code,
                "sortOrder": sort_order,
                "status": status,
                "systemDepartmentCode": system_department_code,
            }
            normalized_departments.append(normalized)
            if department_code:
                department_code_map.setdefault(department_code, []).append(normalized)

        for department_code, items in department_code_map.items():
            if len(items) > 1:
                warnings.append(self._issue("warning", "departments", None, f"업무용 부서 코드 '{department_code}'가 {len(items)}회 중복됩니다."))

        for department in normalized_departments:
            parent_department_name = None
            parent_row_number = None
            if department["parentDepartmentCode"]:
                candidates = department_code_map.get(department["parentDepartmentCode"], [])
                if not candidates:
                    errors.append(self._issue("error", "departments", department["rowNumber"], f"상위 부서 코드 '{department['parentDepartmentCode']}'를 찾지 못했습니다."))
                elif len(candidates) > 1:
                    errors.append(self._issue("error", "departments", department["rowNumber"], f"상위 부서 코드 '{department['parentDepartmentCode']}'가 중복되어 부모 연결이 모호합니다."))
                else:
                    parent_row = candidates[0]
                    if parent_row["rowNumber"] == department["rowNumber"]:
                        errors.append(self._issue("error", "departments", department["rowNumber"], "자기 자신을 상위 부서로 지정할 수 없습니다."))
                    else:
                        parent_row_number = parent_row["rowNumber"]
                        parent_department_name = parent_row["departmentName"]
            department["parentRowNumber"] = parent_row_number
            department["parentDepartmentName"] = parent_department_name

        roles_by_id = {item["id"]: item for item in roles}
        roles_by_name: dict[str, list[dict]] = {}
        for role in roles:
            roles_by_name.setdefault(role["name"], []).append(role)

        existing_users_by_login_id: dict[str, dict] = {}
        for user in existing_users:
            login_id = user["email"].split("@", 1)[0].lower()
            existing_users_by_login_id[login_id] = user

        normalized_users: list[dict] = []
        seen_login_ids: set[str] = set()
        for raw in users:
            row_number = int(raw["rowNumber"])
            login_id = self._normalize_login_id(raw.get("login_id") or raw.get("loginId"))
            name = self._normalize_text(raw.get("name"))
            department_code = self._normalize_text(raw.get("department_code") or raw.get("departmentCode"))
            role_code = self._normalize_text(raw.get("role_code") or raw.get("roleCode"))
            status = self._normalize_status(raw.get("status"), default="active")
            if not login_id:
                errors.append(self._issue("error", "users", row_number, "login_id는 필수이며 영문 소문자/숫자/.-_만 사용할 수 있습니다."))
            elif login_id in seen_login_ids:
                errors.append(self._issue("error", "users", row_number, f"login_id '{login_id}'가 중복됩니다."))
            else:
                seen_login_ids.add(login_id)
            if not name:
                errors.append(self._issue("error", "users", row_number, "name은 필수입니다."))
            if not department_code:
                errors.append(self._issue("error", "users", row_number, "department_code는 필수입니다."))
            role_row = None
            if not role_code:
                errors.append(self._issue("error", "users", row_number, "role_code는 필수입니다."))
            elif role_code in roles_by_id:
                role_row = roles_by_id[role_code]
            else:
                named_roles = roles_by_name.get(role_code, [])
                if not named_roles:
                    errors.append(self._issue("error", "users", row_number, f"권한 역할 '{role_code}'를 찾지 못했습니다."))
                elif len(named_roles) > 1:
                    errors.append(self._issue("error", "users", row_number, f"권한 역할 '{role_code}'가 중복되어 매핑이 모호합니다."))
                else:
                    role_row = named_roles[0]

            department_row = None
            if department_code:
                candidate_departments = department_code_map.get(department_code, [])
                if not candidate_departments:
                    errors.append(self._issue("error", "users", row_number, f"부서 코드 '{department_code}'를 찾지 못했습니다."))
                elif len(candidate_departments) > 1:
                    errors.append(self._issue("error", "users", row_number, f"부서 코드 '{department_code}'가 중복되어 사용자 소속 연결이 모호합니다."))
                else:
                    department_row = candidate_departments[0]

            existing_user = existing_users_by_login_id.get(login_id) if login_id else None
            if existing_user and existing_user["user_type"] == "admin":
                errors.append(self._issue("error", "users", row_number, "관리자 계정은 일괄 업로드 대상으로 사용할 수 없습니다."))

            normalized_users.append(
                {
                    "rowNumber": row_number,
                    "loginId": login_id,
                    "name": name,
                    "departmentCode": department_code,
                    "departmentName": department_row["departmentName"] if department_row else "",
                    "departmentRowNumber": department_row["rowNumber"] if department_row else None,
                    "roleCode": role_code,
                    "roleId": role_row["id"] if role_row else None,
                    "roleName": role_row["name"] if role_row else "",
                    "status": status,
                    "action": "move" if existing_user else "create",
                    "existingUserId": existing_user["id"] if existing_user else None,
                    "existingStatus": existing_user["status"] if existing_user else None,
                }
            )

        normalized_scope = self._normalize_deactivation_scope(deactivation_scope)
        uploaded_login_ids = {item["loginId"] for item in normalized_users if item["loginId"]}
        uploaded_department_codes = {item["departmentCode"] for item in normalized_departments if item["departmentCode"]}
        users_to_deactivate: list[dict] = []
        protected_users: list[dict] = []
        actor_email = str(actor.userEmail).strip().lower() if actor and actor.userEmail else ""

        for user in existing_users:
            email = str(user.get("email") or "").strip().lower()
            login_id = email.split("@", 1)[0] if "@" in email else email
            if user["status"] in {"deleted", "inactive"}:
                continue
            if login_id in uploaded_login_ids:
                continue

            in_uploaded_scope = normalized_scope == "company_all"
            if normalized_scope == "uploaded_departments_only":
                in_uploaded_scope = str(user.get("department_code") or "").strip() in uploaded_department_codes
            if normalized_scope == "none" or not in_uploaded_scope:
                continue

            if self._is_protected_account(user, actor_email):
                protected_users.append({**user, "deactivation_reason": "보호 계정은 업로드 누락 비교에서 제외됩니다."})
                continue

            if normalized_scope == "uploaded_departments_only" and not self._is_verification_account(login_id):
                protected_users.append({**user, "deactivation_reason": "기본 정책에서는 검수용 계정만 자동 비활성화 대상으로 비교합니다."})
                continue

            reason = "회사 전체 누락 비교" if normalized_scope == "company_all" else "업로드 범위 누락"
            users_to_deactivate.append({**user, "deactivation_reason": reason})

        if protected_users:
            warnings.append(
                self._issue(
                    "warning",
                    "users",
                    None,
                    f"보호 계정 {len(protected_users)}명은 비활성화 예정 목록에서 제외되었습니다. 업로드 파일을 수정한 뒤 다시 검토하세요.",
                )
            )
        if normalized_scope == "uploaded_departments_only":
            warnings.append(self._issue("warning", "users", None, "기본 정책은 업로드 부서 범위 안의 검수용 계정만 누락 사용자 비교 대상으로 삼습니다."))
        elif normalized_scope == "company_all":
            warnings.append(self._issue("warning", "users", None, "회사 전체 누락 사용자 비활성화 정책이 선택되었습니다. 보호 계정이 포함되면 적용이 차단됩니다."))

        return {
            "company_id": company["id"],
            "company_domain": company["domain"],
            "current_active_departments": [item for item in existing_departments if item["status"] == "active"],
            "departments": normalized_departments,
            "users": normalized_users,
            "users_to_deactivate": users_to_deactivate,
            "protected_users": protected_users,
            "deactivation_scope": normalized_scope,
            "errors": errors,
            "warnings": warnings,
            "inactive_department_count": len([item for item in existing_departments if item["status"] == "active"]),
            "created_department_count": len(normalized_departments),
            "moved_user_count": len([item for item in normalized_users if item["existingUserId"]]),
            "created_user_count": len([item for item in normalized_users if not item["existingUserId"]]),
            "deactivated_user_count": len(users_to_deactivate),
        }

    def _analysis_to_preview_json(self, analysis: dict) -> dict:
        return {
            "deactivationScope": analysis["deactivation_scope"],
            "departments": [
                {
                    "rowNumber": item["rowNumber"],
                    "systemDepartmentCode": item["systemDepartmentCode"],
                    "departmentCode": item["departmentCode"],
                    "departmentName": item["departmentName"],
                    "parentDepartmentCode": item.get("parentDepartmentCode"),
                    "parentDepartmentName": item.get("parentDepartmentName"),
                    "sortOrder": item["sortOrder"],
                    "status": item["status"],
                }
                for item in analysis["departments"]
            ],
            "users": [
                {
                    "rowNumber": item["rowNumber"],
                    "loginId": item["loginId"],
                    "name": item["name"],
                    "departmentCode": item["departmentCode"],
                    "departmentName": item["departmentName"],
                    "roleCode": item["roleCode"],
                    "roleName": item["roleName"],
                    "status": item["status"],
                    "action": item["action"],
                }
                for item in analysis["users"]
            ],
            "usersToDeactivate": self._deactivation_preview_rows(analysis["users_to_deactivate"]),
            "protectedUsers": self._deactivation_preview_rows(analysis["protected_users"]),
        }

    def _fetch_company_row(self, cursor) -> dict:
        cursor.execute("SELECT id, name, domain, status, created_at FROM companies ORDER BY created_at ASC LIMIT 1")
        row = cursor.fetchone()
        if row is None:
            raise ValueError("초기 설정이 완료되지 않았습니다.")
        return row

    def _fetch_batch_row(self, cursor, batch_id: str, *, for_update: bool = False) -> dict:
        query = "SELECT * FROM org_import_batches WHERE id = %s"
        if for_update:
            query += " FOR UPDATE"
        cursor.execute(query, (batch_id,))
        row = cursor.fetchone()
        if row is None:
            raise ValueError("대상 업로드 배치를 찾을 수 없습니다.")
        return row

    def _row_to_batch_response(self, row: dict) -> OrgImportBatchResponse:
        preview = self._json_value(row["preview_json"], default={})
        return OrgImportBatchResponse(
            batchId=row["id"],
            fileName=row["file_name"],
            uploadedByUserId=row["uploaded_by_user_id"],
            uploadedByUserName=row["uploaded_by_user_name"],
            validationStatus=row["validation_status"],
            applyStatus=row["apply_status"],
            createdDepartmentCount=int(row["created_department_count"]),
            movedUserCount=int(row["moved_user_count"]),
            createdUserCount=int(row["created_user_count"]),
            deactivatedUserCount=int(row["deactivated_user_count"]),
            inactiveDepartmentCount=int(row["inactive_department_count"]),
            errors=[OrgImportIssue(**item) for item in self._json_value(row["errors_json"], default=[])],
            warnings=[OrgImportIssue(**item) for item in self._json_value(row["warnings_json"], default=[])],
            departments=[OrgImportDepartmentPreview(**item) for item in preview.get("departments", [])],
            users=[OrgImportUserPreview(**item) for item in preview.get("users", [])],
            deactivationScope=preview.get("deactivationScope", self.DEFAULT_DEACTIVATION_SCOPE),
            usersToDeactivate=[OrgImportDeactivationPreview(**item) for item in self._json_value(row.get("deactivated_users_json"), default=preview.get("usersToDeactivate", []))],
            protectedUsers=[OrgImportDeactivationPreview(**item) for item in preview.get("protectedUsers", [])],
            uploadedAt=row["uploaded_at"],
            appliedAt=row["applied_at"],
        )


    def _normalize_deactivation_scope(self, scope: str | None) -> str:
        normalized = str(scope or "").strip().lower()
        if normalized in {"none", "uploaded_departments_only", "company_all"}:
            return normalized
        return self.DEFAULT_DEACTIVATION_SCOPE

    def _is_verification_account(self, login_id: str) -> bool:
        normalized = str(login_id or "").strip().lower()
        return any(normalized.startswith(prefix) for prefix in self.VERIFICATION_ACCOUNT_PREFIXES)

    def _is_protected_account(self, user: dict, actor_email: str) -> bool:
        email = str(user.get("email") or "").strip().lower()
        if user.get("user_type") == "admin":
            return True
        if email in self.EXPLICIT_PROTECTED_EMAILS:
            return True
        return bool(actor_email and email == actor_email)

    def _same_preview_targets(self, left: list[dict], right: list[dict]) -> bool:
        left_keys = sorted((str(item.get("email") or "").strip().lower(), str(item.get("reason") or "").strip()) for item in left)
        right_keys = sorted((str(item.get("email") or "").strip().lower(), str(item.get("reason") or "").strip()) for item in right)
        return left_keys == right_keys

    def _deactivation_preview_rows(self, users_to_deactivate: list[dict]) -> list[dict]:
        rows: list[dict] = []
        for user in users_to_deactivate:
            email = str(user.get("email") or "").strip().lower()
            login_id = email.split("@", 1)[0] if "@" in email else email
            rows.append(
                {
                    "userId": user["id"],
                    "loginId": login_id,
                    "name": user["name"],
                    "email": email,
                    "currentDepartmentName": user.get("department_name") or "",
                    "currentRoleName": user.get("role_name") or "",
                    "currentStatus": user.get("status") or "",
                    "reason": str(user.get("deactivation_reason") or "업로드 파일 누락"),
                }
            )
        return rows

    def _insert_audit(
        self,
        *,
        cursor,
        company_id: str,
        actor_user_id: str | None,
        actor_user_name: str,
        target_type: str,
        target_id: str,
        event: str,
        status_before: str | None,
        status_after: str | None,
        reason: str | None,
    ) -> None:
        cursor.execute(
            """
            INSERT INTO audit_logs (
                id, company_id, actor_user_id, actor_user_name, target_type, target_id,
                event, status_before, status_after, reason, created_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                self._new_id("log"),
                company_id,
                actor_user_id,
                actor_user_name,
                target_type,
                target_id,
                event,
                status_before,
                status_after,
                reason,
                self._now(),
            ),
        )

    def _json_value(self, value, *, default):
        if value is None:
            return default
        if isinstance(value, (dict, list)):
            return value
        if isinstance(value, str):
            try:
                return json.loads(value)
            except json.JSONDecodeError:
                return default
        return default

    def _issue(self, level: str, sheet: str, row_number: int | None, message: str) -> dict:
        return {
            "level": level,
            "sheet": sheet,
            "rowNumber": row_number,
            "message": message,
        }

    def _normalize_text(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _normalize_status(self, value, *, default: str) -> str:
        normalized = self._normalize_text(value).lower() or default
        if normalized not in {"active", "inactive"}:
            return default
        return normalized

    def _normalize_int(self, value, *, default: int) -> int:
        text = self._normalize_text(value)
        if not text:
            return default
        try:
            return int(float(text))
        except ValueError:
            return default

    def _normalize_login_id(self, value) -> str:
        normalized = self._normalize_text(value).lower()
        if not normalized:
            return ""
        allowed = set("abcdefghijklmnopqrstuvwxyz0123456789._-")
        if any(char not in allowed for char in normalized):
            return ""
        return normalized

    def _new_unique_system_department_code(self, used_codes: set[str]) -> str:
        while True:
            candidate = f"DPT-{uuid4().hex[:10].upper()}"
            if candidate not in used_codes:
                used_codes.add(candidate)
                return candidate

    def _new_id(self, prefix: str) -> str:
        return f"{prefix}_{uuid4().hex[:12]}"

    def _now(self) -> datetime:
        return datetime.now(UTC)
